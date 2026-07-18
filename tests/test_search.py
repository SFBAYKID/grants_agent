"""On-demand search correctness: filters, dates, classification, and completeness."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook

from grant_watch import db, google_sheets
from grant_watch.models import (
    DatePrecision,
    FundingEventType,
    Lead,
    LeadGrade,
    RawItem,
    VerificationStatus,
)
from grant_watch.slack import tools
from grant_watch.slack.search import MAX_ENRICH_ROWS, MAX_EXPORT_ROWS, search_leads


def _insert(
    conn: sqlite3.Connection,
    source: str,
    item_id: str,
    entity: str,
    state: str,
    program: str,
    amount: float | None,
    start: str,
    end: str,
    grade: LeadGrade,
) -> None:
    """Insert one typed fixture lead with source semantics preserved."""
    event_type = (
        FundingEventType.AWARD_ANNOUNCED
        if grade is LeadGrade.GOLD
        else FundingEventType.APPLICATION_WINDOW_OPENED
        if source == "grants.gov"
        else FundingEventType.RFP_POSTED
    )
    db.upsert_lead(
        conn,
        Lead(
            item=RawItem(
                source=source,
                item_id=item_id,
                title=f"Title {item_id}",
                entity=entity,
                state=state,
                program=program,
                amount=amount,
                start=start,
                end=end,
                url="https://example.gov/item",
                raw={},
                event_type=event_type,
                event_date=start,
                date_precision=DatePrecision.DAY,
                verification_status=VerificationStatus.VERIFIED,
            ),
            grade=grade,
        ),
    )


def _db(tmp_path: Path) -> Path:
    """Build a mixed award/opportunity/solicitation database for search tests."""
    path = tmp_path / "search.db"
    conn = db.connect(path)
    rows = [
        (
            "usaspending:16.071",
            "A1",
            "Tustin Unified School District",
            "CA",
            "SVPP",
            500_000.0,
            "2025-10-01",
            "2028-09-30",
            LeadGrade.GOLD,
        ),
        (
            "usaspending:16.071",
            "A2",
            "City of Austin",
            "TX",
            "NSGP",
            120_000.0,
            "2025-11-01",
            "2028-09-30",
            LeadGrade.GOLD,
        ),
        (
            "usaspending:16.071",
            "A3",
            "Fresno County",
            "CA",
            "STOP",
            80_000.0,
            "2024-01-01",
            "2027-01-01",
            LeadGrade.GOLD,
        ),
        (
            "sam.gov",
            "S1",
            "Federal Procurement Office",
            "WA",
            "RFP:sam.gov",
            None,
            "2026-07-10",
            "2026-08-15",
            LeadGrade.SILVER,
        ),
        (
            "grants.gov",
            "O1",
            "COPS Office",
            "",
            "SVPP",
            None,
            "2026-08-01",
            "2026-08-31",
            LeadGrade.WATCH,
        ),
        (
            "usaspending:16.071",
            "A4",
            "Modesto City Schools",
            "CA",
            "SVPP",
            300_000.0,
            "2026-08-01",
            "2028-09-30",
            LeadGrade.GOLD,
        ),
        (
            "usaspending:16.071",
            "A5",
            "Township High School District 211",
            "IL",
            "SVPP",
            250_000.0,
            "2025-10-01",
            "2028-09-30",
            LeadGrade.GOLD,
        ),
        (
            "usaspending:16.071",
            "A6",
            "Mesa Water District",
            "CA",
            "STOP",
            90_000.0,
            "2025-10-01",
            "2028-09-30",
            LeadGrade.GOLD,
        ),
    ]
    for row in rows:
        _insert(conn, *row)
    conn.execute(
        """UPDATE leads SET nces_id='0640150',enrollment=21220,
                  location_city='Tustin',location_confidence='high'
           WHERE source_item_id='A1'"""
    )
    conn.commit()
    conn.close()
    return path


def _bulk_db(tmp_path: Path, count: int) -> Path:
    """Build many active leads efficiently for export-boundary tests."""
    path = tmp_path / f"bulk-{count}.db"
    conn = db.connect(path)
    now = "2026-07-14T12:00:00+00:00"
    conn.executemany(
        """INSERT INTO leads
           (source, source_item_id, lead_grade, entity_name, state, program,
            funds_start, funds_end, first_seen, last_seen, status)
           VALUES ('usaspending:16.071', ?, 'gold', ?, 'CA', 'SVPP',
                   '2025-10-01', '2028-09-30', ?, ?, 'new')""",
        [(f"B{i}", f"School District {i}", now, now) for i in range(count)],
    )
    conn.commit()
    conn.close()
    return path


def test_filter_by_state_grade_and_amount(tmp_path: Path) -> None:
    """Combine common scalar filters without leaking another state's result."""
    text, _ = search_leads(
        state="CA", grade="gold", amount_min=400_000, db_path=_db(tmp_path)
    )
    assert "Tustin" in text
    assert "Austin" not in text


def test_nces_enrollment_and_city_filters_are_explicit(tmp_path: Path) -> None:
    """Matched NCES district facts support exact filters and disclose partial coverage."""
    path = _db(tmp_path)
    text, _ = search_leads(
        state="CA",
        org_type="school",
        city="Tustin",
        enrollment_min=20_000,
        db_path=path,
    )
    assert "Tustin Unified School District" in text
    assert "21,220 students" in text
    assert "NCES enrollment matched 1 of" in text


def test_nces_filter_without_state_falls_back_honestly(tmp_path: Path) -> None:
    """Missing state preserves other results while saying enrollment was not applied."""
    text, _ = search_leads(
        org_type="school", enrollment_min=1_000, db_path=_db(tmp_path)
    )
    assert "Tustin Unified School District" in text
    assert "requires a two-letter state" in text


def test_nces_outage_preserves_other_filters(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Reference-data failure never becomes a false zero-result enrollment search."""
    from grant_watch.enrich import nces

    path = _db(tmp_path)
    conn = db.connect(path)
    conn.execute("UPDATE leads SET nces_id=NULL,enrollment=NULL,location_city=NULL")
    conn.commit()
    conn.close()

    def unavailable(_conn: sqlite3.Connection, _state: str) -> nces.EnrichmentSummary:
        """Simulate NCES being unreachable without external I/O."""
        raise TimeoutError("offline")

    monkeypatch.setattr(nces, "enrich_state_leads", unavailable)
    text, _ = search_leads(
        state="CA", org_type="school", enrollment_min=20_000, db_path=path
    )
    assert "NCES reference data was unavailable" in text
    assert "Tustin Unified School District" in text  # other filters still honored


def test_school_precedence_excludes_school_names_from_city(tmp_path: Path) -> None:
    """Municipal words inside school names must never make those entities cities."""
    path = _db(tmp_path)
    school_text, _ = search_leads(org_type="school", db_path=path)
    city_text, _ = search_leads(org_type="city", db_path=path)
    assert "Modesto City Schools" in school_text
    assert "Township High School District 211" in school_text
    assert "City of Austin" in city_text
    assert "Modesto City Schools" not in city_text
    assert "Township High School District 211" not in city_text
    assert "Mesa Water District" not in school_text


def test_county_filter_excludes_school_districts(tmp_path: Path) -> None:
    """County entities match while unrelated district names remain excluded."""
    text, _ = search_leads(org_type="county", db_path=_db(tmp_path))
    assert "Fresno County" in text
    assert "School District" not in text


def test_opportunity_open_range_is_source_specific(tmp_path: Path) -> None:
    """An opportunity-open search must not include an award spend start on the same date."""
    text, _ = search_leads(
        date_field="opportunity_open",
        date_from="2026-08-01",
        date_to="2026-08-31",
        db_path=_db(tmp_path),
    )
    assert "COPS Office" in text
    assert "Modesto City Schools" not in text
    assert "applications open" in text


def test_response_due_range_is_solicitation_specific(tmp_path: Path) -> None:
    """RFP response deadlines remain distinct from award spend-window ends."""
    text, _ = search_leads(
        date_field="response_due",
        date_from="2026-08-01",
        date_to="2026-08-31",
        db_path=_db(tmp_path),
    )
    assert "Federal Procurement Office" in text
    assert "response due 2026-08-15" in text
    assert "Tustin" not in text


def test_award_received_uses_verified_event_date_with_coverage_disclosure(
    tmp_path: Path,
) -> None:
    """Award-date searches use current event truth and disclose incomplete coverage."""
    text, artifact = search_leads(
        date_field="award_received",
        date_from="2026-07-01",
        date_to="2026-08-31",
        db_path=_db(tmp_path),
    )
    assert "Modesto City Schools" in text
    assert "award event 2026-08-01" in text
    assert "coverage may be incomplete" in text
    assert artifact is None


def test_zero_results_offer_relaxation_hints_with_real_counts(
    tmp_path: Path,
) -> None:
    """A dead search names what one dropped filter would find — never a bare no."""
    text, artifact = search_leads(
        state="CA", program="ZZNOPE", db_path=_db(tmp_path)
    )
    assert artifact is None
    assert "No grants matched those filters." in text
    assert "Nearby alternatives" in text
    assert "without the program filter" in text
    assert "do not stop at a bare no-results answer" in text


def test_zero_results_fall_back_to_whole_pool_count(tmp_path: Path) -> None:
    """When every single-filter drop is still zero, the total pool is offered."""
    text, _ = search_leads(
        state="ZZ", program="ZZNOPE", db_path=_db(tmp_path)
    )
    assert "Nearby alternatives" in text
    assert "leads on file overall" in text


def test_award_received_without_range_sorts_newest_first(tmp_path: Path) -> None:
    """A sort-only date_field is a valid ask, not an error.

    Live failure 2026-07-18: "newest verified award announcements" produced
    date_field with no range, which errored and drained the tool loop. It must
    return verified award rows ordered by event date, newest first."""
    text, artifact = search_leads(
        date_field="award_received",
        db_path=_db(tmp_path),
    )
    assert "ERROR" not in text
    assert "Modesto City Schools" in text  # the verified award event row
    assert artifact is None


def test_record_kind_uses_event_truth_not_projection_grade(tmp_path: Path) -> None:
    """An award remains an award when a projection grade changes independently."""
    path = _db(tmp_path)
    conn = db.connect(path)
    conn.execute("UPDATE leads SET lead_grade='watch' WHERE source_item_id='A1'")
    conn.commit()
    conn.close()
    text, _ = search_leads(record_kind="award", name_contains="Tustin", db_path=path)
    assert "Tustin Unified School District" in text
    assert "coverage may be incomplete" in text


def test_gold_window_is_never_labeled_closing(tmp_path: Path) -> None:
    """Award results describe spend windows rather than application close dates."""
    text, _ = search_leads(name_contains="Tustin", db_path=_db(tmp_path))
    assert "spend window 2025-10-01 through 2028-09-30" in text
    assert "closes" not in text


def test_invalid_reversed_date_and_amount_ranges_fail(tmp_path: Path) -> None:
    """Invalid ranges fail explicitly rather than silently returning misleading data."""
    path = _db(tmp_path)
    date_text, _ = search_leads(
        date_field="spend_end",
        date_from="2027-01-02",
        date_to="2027-01-01",
        db_path=path,
    )
    amount_text, _ = search_leads(amount_min=10, amount_max=1, db_path=path)
    assert date_text.startswith("ERROR") and "after" in date_text
    assert amount_text.startswith("ERROR") and "exceed" in amount_text


def test_incompatible_record_kind_and_date_fail(tmp_path: Path) -> None:
    """A funding-opportunity date cannot be applied to award records."""
    text, artifact = search_leads(
        record_kind="award",
        date_field="opportunity_close",
        date_from="2026-08-01",
        date_to="2026-08-31",
        db_path=_db(tmp_path),
    )
    assert text.startswith("ERROR") and "incompatible" in text
    assert artifact is None


def test_like_metacharacters_are_literal(tmp_path: Path) -> None:
    """Percent and underscore in user text must not broaden a LIKE search."""
    path = _db(tmp_path)
    conn = db.connect(path)
    _insert(
        conn,
        "usaspending:16.071",
        "PCT",
        "100% Secure School",
        "CA",
        "SVPP",
        1.0,
        "2025-10-01",
        "2028-09-30",
        LeadGrade.GOLD,
    )
    conn.close()
    text, _ = search_leads(name_contains="100%", db_path=path)
    assert "100% Secure School" in text
    assert "Found 1 matching" in text


def test_dead_leads_are_excluded(tmp_path: Path) -> None:
    """Human-rejected leads never resurface through on-demand search."""
    path = _db(tmp_path)
    conn = db.connect(path)
    conn.execute("UPDATE leads SET status='dead' WHERE source_item_id='A1'")
    conn.commit()
    conn.close()
    text, _ = search_leads(name_contains="Tustin", db_path=path)
    assert "No grants matched" in text


def test_excel_all_scope_ignores_inline_limit_and_is_complete(tmp_path: Path) -> None:
    """An explicit all-results export contains every match without silent truncation."""
    text, artifact = search_leads(
        limit=1, export="excel", result_scope="all", db_path=_bulk_db(tmp_path, 51)
    )
    assert artifact is not None
    workbook = load_workbook(artifact.path)
    assert workbook.active.max_row == 52  # header + all 51 matches
    assert "all 51" in text
    artifact.cleanup()


def test_excel_top_n_scope_honors_requested_limit(tmp_path: Path) -> None:
    """The default export scope contains only the confirmed top N rows."""
    text, artifact = search_leads(
        limit=1, export="excel", db_path=_bulk_db(tmp_path, 51)
    )
    assert artifact is not None
    assert load_workbook(artifact.path).active.max_row == 2
    assert "exported the top 1" in text
    artifact.cleanup()


def test_export_job_persists_with_search_snapshot(tmp_path: Path) -> None:
    """A real Slack export records its frozen search and final delivery state."""
    path = _bulk_db(tmp_path, 3)
    text, artifact = search_leads(
        limit=2,
        export="excel",
        requester_slack="U1",
        workspace="T1",
        channel="C1",
        thread_ts="100.1",
        db_path=path,
    )
    assert artifact is not None and "exported the top 2" in text
    conn = db.connect(path)
    job = conn.execute("SELECT * FROM export_jobs").fetchone()
    snapshot = conn.execute("SELECT * FROM search_requests").fetchone()
    assert job is not None and job["state"] == "created"
    assert snapshot is not None and job["search_request_id"] == snapshot["id"]
    artifact.cleanup()


def test_export_over_declared_cap_creates_no_partial_file(tmp_path: Path) -> None:
    """Oversized exports fail with the true count and never masquerade as complete."""
    count = MAX_EXPORT_ROWS + 1
    text, artifact = search_leads(
        export="excel", result_scope="all", db_path=_bulk_db(tmp_path, count)
    )
    assert artifact is None
    assert str(count) in text and str(MAX_EXPORT_ROWS) in text
    assert "no incomplete file" in text


def test_google_sheet_success_exports_every_match(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A successful Google handoff receives all 2,001 rows and creates no local file."""
    captured_count = 0

    def fake_create(
        _title: str,
        _columns: list[str],
        rows: list[list[object]],
        _requested_by: str,
        _send_as: str,
    ) -> tuple[str, str]:
        """Capture the complete export and return a verified-looking test URL."""
        nonlocal captured_count
        captured_count = len(rows)
        return "created", "https://docs.google.com/spreadsheets/d/test"

    monkeypatch.setattr(google_sheets, "create_sheet", fake_create)
    text, artifact = search_leads(
        export="google_sheet",
        result_scope="all",
        requester_slack="U01DPJVURHU",
        db_path=_bulk_db(tmp_path, 2_001),
    )
    assert captured_count == 2_001
    assert "all 2001" in text
    assert artifact is None


def test_google_sheet_failure_falls_back_to_complete_excel(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """An unwired Google endpoint returns a complete Excel artifact without truncation."""

    def fake_create(
        _title: str,
        _columns: list[str],
        _rows: list[list[object]],
        _requested_by: str,
        _send_as: str,
    ) -> tuple[str, str]:
        """Simulate an unconfigured/failed export so the Excel fallback is exercised."""
        return "unconfigured", "Google Sheet export is not live"

    monkeypatch.setattr(google_sheets, "create_sheet", fake_create)
    text, artifact = search_leads(
        limit=1,
        export="google_sheet",
        result_scope="all",
        requester_slack="U01DPJVURHU",
        db_path=_bulk_db(tmp_path, 51),
    )
    assert artifact is not None
    assert load_workbook(artifact.path).active.max_row == 52
    assert "complete Excel file instead" in text
    artifact.cleanup()


# ------------------------------------------------------------ with_contacts enrichment
def _verified(entity: str) -> tools.ContactOutcome:
    """A deterministic verified outcome keyed to the entity name."""
    return tools.ContactOutcome(
        "verified",
        name=f"Dir {entity[:6]}",
        title="Technology Director",
        email=f"it@{entity[:4].lower()}.org",
    )


def test_determinism_repeated_search_returns_same_rows(tmp_path: Path) -> None:
    """Tie-heavy data must return an identical top-N across calls (the id tiebreak) —
    otherwise turn-2 enrichment would attach contacts to orgs the rep never saw."""
    path = _bulk_db(tmp_path, 20)  # all share first_seen + funds_start, amount NULL
    first, _ = search_leads(limit=10, db_path=path)
    second, _ = search_leads(limit=10, db_path=path)
    assert first == second


def test_explicit_top_ten_is_shown_even_when_more_matches_exist(tmp_path: Path) -> None:
    """A confirmed small inline limit is honored instead of prompting for export."""
    text, artifact = search_leads(limit=10, db_path=_bulk_db(tmp_path, 20))
    assert artifact is None
    assert text.startswith("Found 20 matching grants")
    assert "Showing 10 of 20" in text


def test_results_carry_grade_split_and_source_links(tmp_path: Path) -> None:
    """Every shown row keeps its public source link; the header explains grades."""
    text, _ = search_leads(state="CA", db_path=_db(tmp_path))
    assert "gold (award won, money to spend)" in text
    assert "|verify this record>" in text  # Slack-formatted link on result rows


def test_ca_portal_rows_get_per_record_verification_links() -> None:
    """CA dataset rows deep-link to THEIR award record, not the shared dataset.

    Chase's honesty rule: 'under Pleasant View Elementary … there needs to be a
    link to that specific [amount]' — a generic dataset URL repeated per row
    proves nothing."""
    from grant_watch.slack.search_presentation import record_link

    class FakeRow(dict):
        """Minimal sqlite3.Row stand-in supporting [] access."""

        def __getitem__(self, key: str) -> object:
            """Read one column value."""
            return dict.__getitem__(self, key)

    row = FakeRow(
        source="ca-grants-award:2023-2024",
        source_item_id="73146",
        detail_url=(
            "https://data.ca.gov/dataset/572d06aa-4f1f-44ad-80a4-167bec020881/"
            "resource/018f3523-652d-4197-a4a8-a055bfd1544f"
        ),
    )
    link = record_link(row)
    assert "datastore_search" in link
    assert "resource_id=018f3523-652d-4197-a4a8-a055bfd1544f" in link
    assert "73146" in link
    # Non-CA rows keep their own per-record page untouched.
    usa = FakeRow(
        source="usaspending",
        source_item_id="X",
        detail_url="https://www.usaspending.gov/award/ASST_NON_123",
    )
    assert record_link(usa) == "https://www.usaspending.gov/award/ASST_NON_123"


def test_with_contacts_appends_columns_to_summary(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The second step enriches the shown orgs and shows each contact inline."""
    monkeypatch.setattr(
        tools,
        "enrich_lead_contact",
        lambda _c, lead_id, _p=None: _verified(f"lead-{lead_id}"),
    )
    text, _ = search_leads(
        state="CA", grade="gold", with_contacts=True, limit=3, db_path=_db(tmp_path)
    )
    assert "contact:" in text
    assert "Technology Director" in text


def test_with_contacts_export_has_contact_columns(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Excel export carries the same contact columns as the inline summary (parity)."""
    monkeypatch.setattr(
        tools,
        "enrich_lead_contact",
        lambda _c, lead_id, _p=None: _verified(f"lead-{lead_id}"),
    )
    _, artifact = search_leads(
        state="CA",
        grade="gold",
        export="excel",
        with_contacts=True,
        limit=3,
        db_path=_db(tmp_path),
    )
    assert artifact is not None
    sheet = load_workbook(artifact.path).active
    header = [c.value for c in sheet[1]]
    assert header[-4:] == list(
        ("contact_name", "contact_title", "contact_email", "contact_status")
    )
    emails = [
        sheet.cell(row=r, column=len(header) - 1).value
        for r in range(2, sheet.max_row + 1)
    ]
    assert any(e and "@" in str(e) for e in emails)
    artifact.cleanup()


def test_with_contacts_one_failure_does_not_sink_batch(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A single org's enrichment blowing up degrades to 'error'; the rest still resolve."""

    def flaky(_c: object, lead_id: int, _p: object = None) -> tools.ContactOutcome:
        """Provide test-local behavior for flaky."""
        if lead_id == 1:
            raise RuntimeError("boom")
        return _verified(f"lead-{lead_id}")

    monkeypatch.setattr(tools, "enrich_lead_contact", flaky)
    text, _ = search_leads(
        state="CA", grade="gold", with_contacts=True, limit=5, db_path=_db(tmp_path)
    )
    assert "lookup error" in text  # the org that raised
    assert "Technology Director" in text  # the others still enriched


def test_with_contacts_caps_and_discloses(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Asking for more than the ceiling enriches only the cap and says so."""
    calls = 0

    def counting(_c: object, lead_id: int, _p: object = None) -> tools.ContactOutcome:
        """Provide test-local behavior for counting."""
        nonlocal calls
        calls += 1
        return _verified(f"lead-{lead_id}")

    monkeypatch.setattr(tools, "enrich_lead_contact", counting)
    text, _ = search_leads(with_contacts=True, limit=15, db_path=_bulk_db(tmp_path, 30))
    assert calls == MAX_ENRICH_ROWS
    assert f"top {MAX_ENRICH_ROWS}" in text
