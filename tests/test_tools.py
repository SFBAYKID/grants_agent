"""Grant's tool layer: safety guards and real outputs (all offline)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from grant_watch import db
from grant_watch.models import Lead, LeadGrade, RawItem
from grant_watch.slack import tools
from grant_watch.spreadsheets import make_spreadsheet


# ------------------------------------------------------------------ lead_stats
def test_lead_stats_uses_typed_allowlisted_view(tmp_path: Path) -> None:
    """Counts come from a throwaway DB without model-authored SQL or local state."""
    path = tmp_path / "stats.db"
    conn = db.connect(path)
    db.upsert_lead(
        conn,
        Lead(
            RawItem(
                "test",
                "1",
                "award",
                "Alpha District",
                "CA",
                "SVPP",
                100_000,
                "2026-01-01",
                "2027-01-01",
                "",
                {},
            ),
            LeadGrade.GOLD,
        ),
    )
    out = tools.lead_stats(group_by="grade", db_path=path)
    assert "gold: 1" in out


def test_raw_sql_tool_is_not_exposed() -> None:
    """The conversational model has no arbitrary database query capability."""
    names = {schema["name"] for schema in tools.TOOL_SCHEMAS}
    assert "query_leads" not in names
    assert "make_spreadsheet" not in names
    assert "lead_stats" in names
    search_schema = next(
        schema for schema in tools.TOOL_SCHEMAS if schema["name"] == "search_leads"
    )
    properties = search_schema["input_schema"]["properties"]
    assert {"city", "enrollment_min", "enrollment_max"} <= set(properties)


# ------------------------------------------------------------------ spreadsheet
def test_make_spreadsheet_builds_real_xlsx() -> None:
    """The spreadsheet helper creates a readable workbook with numeric cells intact."""
    text, artifact = make_spreadsheet(
        "test export.xlsx", [["entity", "amount"], ["Castle Rock SD", 500000]]
    )
    assert "attached" in text
    wb = load_workbook(artifact.path)
    ws = wb.active
    assert ws["A1"].value == "entity"
    assert ws["B2"].value == 500000
    artifact.cleanup()


def test_make_spreadsheet_sanitizes_filename() -> None:
    """User-supplied filenames cannot escape the private temporary directory."""
    _, artifact = make_spreadsheet("../..//evil name!!", [["a"]])
    name = artifact.path.name
    assert "/" not in name and name.endswith(".xlsx")
    artifact.cleanup()


def test_make_spreadsheet_neutralizes_formulas_but_keeps_negative_numbers() -> None:
    """External formula-like strings become text while genuine numbers remain numeric."""
    _, artifact = make_spreadsheet(
        "safe.xlsx",
        [
            ["equals", "plus", "minus", "at", "spaced", "negative"],
            ["=1+1", "+cmd", "-formula", "@name", " \t=SUM(A1:A2)", -25.0],
        ],
    )
    sheet = load_workbook(artifact.path, data_only=False).active
    for column in ("A", "B", "C", "D", "E"):
        assert sheet[f"{column}2"].data_type == "s"
        assert str(sheet[f"{column}2"].value).startswith("'")
    assert sheet["F2"].value == -25.0
    assert sheet["F2"].data_type == "n"
    artifact.cleanup()


# ------------------------------------------------------------------ dispatch
def test_run_tool_unknown_is_honest_error() -> None:
    """Unknown tool names fail explicitly without creating an artifact."""
    text, artifact = tools.run_tool("teleport", {})
    assert text.startswith("ERROR") and artifact is None


# ------------------------------------------------------------------ name resolution
def _seed(conn: "db.sqlite3.Connection", item_id: str, entity: str, state: str) -> int:
    """Insert one lead and return its id."""
    db.upsert_lead(
        conn,
        Lead(
            RawItem(
                "test", item_id, "award", entity, state, "SVPP",
                100_000, "2026-01-01", "2027-01-01", "", {},
            ),
            LeadGrade.GOLD,
        ),
    )
    return int(
        conn.execute(
            "SELECT id FROM leads WHERE source='test' AND source_item_id=?",
            (item_id,),
        ).fetchone()[0]
    )


def test_resolve_lead_by_name_unique_match(tmp_path: Path) -> None:
    """An exact organization name resolves to its single lead id."""
    conn = db.connect(tmp_path / "r.db")
    lead_id = _seed(conn, "r1", "Chicago Jewish Day School", "IL")
    assert tools.resolve_lead_by_name(conn, "Chicago Jewish Day School") == lead_id


def test_resolve_lead_by_name_state_disambiguates(tmp_path: Path) -> None:
    """Two same-named orgs in different states resolve only with the state."""
    conn = db.connect(tmp_path / "r.db")
    il_id = _seed(conn, "r2", "Lincoln Elementary", "IL")
    _seed(conn, "r3", "Lincoln Elementary", "CA")
    ambiguous = tools.resolve_lead_by_name(conn, "Lincoln Elementary")
    assert isinstance(ambiguous, str) and "several Grant leads" in ambiguous
    assert tools.resolve_lead_by_name(conn, "Lincoln Elementary", "IL") == il_id


def test_resolve_lead_by_name_unknown_is_honest(tmp_path: Path) -> None:
    """A name with no lead returns an honest error, never a guess."""
    conn = db.connect(tmp_path / "r.db")
    out = tools.resolve_lead_by_name(conn, "Nonexistent Academy")
    assert isinstance(out, str) and "no Grant lead is named" in out


# ------------------------------------------------- contact-record preview binding
def test_contact_preview_resolves_lead_by_entity_name(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """"add <person> to Salesforce" resolves the lead from the org name — a natural
    request never carries a lead number (live snag 2026-07-18: City of East
    Providence dead-ended because the preview tool demanded a lead_id)."""
    from grant_watch.enrich import salesforce_campaign_gateway as gw
    from grant_watch.enrich import salesforce_contact_records as records

    conn = db.connect(tmp_path / "p.db")
    lead_id = _seed(conn, "p1", "City of East Providence", "RI")
    monkeypatch.setattr(tools.db, "connect", lambda *a, **k: conn)
    monkeypatch.setattr(gw, "SalesforceCampaignGateway", lambda *a, **k: object())
    captured: dict[str, int] = {}

    class _Prep:
        """Minimal PreparedAction stand-in."""

        action_id = "a1"
        nonce = "n1"
        preview = "PREVIEW-TEXT"
        expires_at = "2026-07-18T23:00:00+00:00"

    def _fake_prepare(
        _conn: object,
        _gateway: object,
        _ws: str,
        _ch: str,
        _ts: str,
        _req: str,
        lid: int,
        cid: int | None = None,
    ) -> _Prep:
        """Capture the lead_id the tool resolved."""
        captured["lead_id"] = lid
        return _Prep()

    monkeypatch.setattr(records, "prepare_contact_record", _fake_prepare)
    out = tools.salesforce_contact_record_preview(
        {"entity": "City of East Providence", "state": "RI"}, "U1", "W", "C", "T"
    )
    assert captured["lead_id"] == lead_id  # resolved by name, never asked for
    assert not out.startswith("ERROR")


def test_contact_preview_unresolved_entity_is_honest(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """An org name with no lead returns the honest resolver error, no write attempt."""
    conn = db.connect(tmp_path / "p.db")  # empty
    monkeypatch.setattr(tools.db, "connect", lambda *a, **k: conn)
    out = tools.salesforce_contact_record_preview(
        {"entity": "Nonexistent City"}, "U1", "W", "C", "T"
    )
    assert out.startswith("ERROR") and "no Grant lead is named" in out
